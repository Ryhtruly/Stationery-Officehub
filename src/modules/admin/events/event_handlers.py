import matplotlib
matplotlib.use('Qt5Agg')

from datetime import datetime
from PyQt5 import QtWidgets
from PyQt5.QtCore import QDate
from PyQt5.QtGui import QFont, QColor
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QTableWidgetItem
from src.modules.admin.data.data_loader import StatisticDataLoader
import pandas as pd
import os

statistic_data_loader = StatisticDataLoader()

def update_statistics_event(window):
    """Cập nhật dữ liệu thống kê"""
    try:
        # Lấy ngày từ UI
        from_date = window.from_date.date().toString("dd-MM-yyyy")
        to_date = window.to_date.date().toString("dd-MM-yyyy")

        update_chart(window, from_date, to_date)

        data, total_quantity, total_revenue, total_cost, total_profit = statistic_data_loader.get_product_statistics(
            from_date, to_date)

        print(f"Debug - Tổng số lượng: {total_quantity}")
        print(f"Debug - Tổng doanh thu: {total_revenue}")
        print(f"Debug - Tổng vốn: {total_cost}")
        print(f"Debug - Tổng lợi nhuận: {total_profit}")

        statistic_data_loader.load_data_to_statistic_table(window.table, from_date, to_date)

        row_count = window.table.rowCount()
        window.table.insertRow(row_count)
        window.table.setItem(row_count, 0, QTableWidgetItem(""))
        window.table.setItem(row_count, 1, QTableWidgetItem("TỔNG"))
        window.table.setItem(row_count, 2, QTableWidgetItem(str(total_quantity)))
        window.table.setItem(row_count, 3, QTableWidgetItem(f"{float(total_revenue):,.0f}"))
        window.table.setItem(row_count, 4, QTableWidgetItem(f"{float(total_cost):,.0f}"))
        window.table.setItem(row_count, 5, QTableWidgetItem(f"{float(total_profit):,.0f}"))

        for col in range(6):
            item = window.table.item(row_count, col)
            if item:
                item.setBackground(QColor("#e9c46a"))
                item.setFont(QFont("Arial", 10, QFont.Bold))

        # Cập nhật các label tổng (giả sử bạn có các label này)
        if hasattr(window, 'total_quantity_label'):
            window.total_quantity_label.setText(f"Tổng số lượng: {total_quantity}")

        if hasattr(window, 'total_revenue_label'):
            window.total_revenue_label.setText(f"Tổng doanh thu: {float(total_revenue):,.0f} VNĐ")

        if hasattr(window, 'total_cost_label'):
            window.total_cost_label.setText(f"Tổng vốn: {float(total_cost):,.0f} VNĐ")

        if hasattr(window, 'total_profit_label'):
            window.total_profit_label.setText(f"Tổng lợi nhuận: {float(total_profit):,.0f} VNĐ")

    except Exception as e:
        print(f"Lỗi chi tiết: {str(e)}")
        import traceback
        traceback.print_exc()
        QMessageBox.warning(window, "Lỗi", f"Không thể cập nhật thống kê: {str(e)}")

def update_chart(window, from_date, to_date):
    try:
        # Đảm bảo from_date và to_date là chuỗi
        if isinstance(from_date, QDate):
            from_date = from_date.toString("dd-MM-yyyy")
        if isinstance(to_date, QDate):
            to_date = to_date.toString("dd-MM-yyyy")

        print(f"From date: {from_date}, To date: {to_date}")

        # Lấy dữ liệu thống kê theo ngày
        dates, revenues = statistic_data_loader.get_daily_statistics(from_date, to_date)

        print(f"Dữ liệu biểu đồ: {dates}")

        # Xóa biểu đồ cũ
        window.figure.clear()

        # Tạo biểu đồ mới nếu có dữ liệu
        if dates and revenues:
            ax = window.figure.add_subplot(111)

            # Xác định loại biểu đồ
            chart_type = window.chart_type_combo.currentText() if hasattr(window, 'chart_type_combo') else "Cột"

            if chart_type == "Cột":
                ax.bar(dates, revenues, color='#2196F3')
            elif chart_type == "Đường":
                ax.plot(dates, revenues, 'o-', color='#2196F3', linewidth=2)
            elif chart_type == "Vùng":
                ax.fill_between(dates, revenues, color='#2196F3', alpha=0.5)
                ax.plot(dates, revenues, color='#0b7dda', linewidth=2)

            ax.set_title('Doanh thu theo ngày')
            ax.set_xlabel('Ngày')
            ax.set_ylabel('Doanh thu (VNĐ)')

            # Định dạng trục x để hiển thị ngày rõ ràng hơn
            if len(dates) > 10:
                # Nếu có nhiều ngày, chỉ hiển thị một số ngày để tránh chồng chéo
                step = max(1, len(dates) // 10)
                ax.set_xticks(range(0, len(dates), step))
                ax.set_xticklabels([dates[i] for i in range(0, len(dates), step)])

            ax.tick_params(axis='x', rotation=45)

            # Định dạng số tiền trên trục y
            import matplotlib.ticker as ticker
            ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, pos: f'{x:,.0f}'))

            window.figure.tight_layout()
        else:
            # Hiển thị thông báo nếu không có dữ liệu
            ax = window.figure.add_subplot(111)
            ax.text(0.5, 0.5, 'Không có dữ liệu trong khoảng thời gian này',
                    horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes, fontsize=12)
            ax.axis('off')

        # Cập nhật canvas
        window.canvas.draw()

    except Exception as e:
        print(f"Lỗi khi cập nhật biểu đồ: {str(e)}")

def export_report_event(window, openpyxl=None):
    """Xuất báo cáo thống kê"""
    try:
        from_date = window.from_date.date().toString("dd-MM-yyyy")
        to_date = window.to_date.date().toString("dd-MM-yyyy")

        # Lấy dữ liệu thống kê
        data, total_quantity, total_revenue, total_cost, total_profit = statistic_data_loader.get_product_statistics(
            from_date, to_date)

        if not data:
            QMessageBox.warning(window, "Thông báo", "Không có dữ liệu để xuất báo cáo!")
            return

        # Mở hộp thoại lưu file
        file_path, _ = QFileDialog.getSaveFileName(
            window,
            "Lưu báo cáo",
            f"Thong_ke_{from_date}_den_{to_date}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        # Tạo DataFrame từ dữ liệu
        import pandas as pd

        df = pd.DataFrame(data, columns=['ID', 'Tên sản phẩm', 'Số lượng', 'Tổng thu nhập', 'Tổng vốn', 'Lợi nhuận'])

        # Thêm dòng tổng
        df.loc[len(df)] = ['', 'TỔNG', total_quantity, total_revenue, total_cost, total_profit]

        # Xuất ra file Excel
        df.to_excel(file_path, index=False, sheet_name='Thống kê sản phẩm')

        # Định dạng file Excel
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = load_workbook(file_path)
        ws = wb.active

        # Định dạng tiêu đề
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Định dạng dòng tổng
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.font = Font(bold=True)

        # Lưu lại file
        wb.save(file_path)

        QMessageBox.information(window, "Thông báo", f"Đã xuất báo cáo thành công!\nĐường dẫn: {file_path}")

    except Exception as e:
        QMessageBox.critical(window, "Lỗi", f"Không thể xuất báo cáo: {str(e)}")
